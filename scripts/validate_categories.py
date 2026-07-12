#!/usr/bin/env python3
"""
產品分類驗證器 — 防止分類汙染再次發生。

用品名規則檢查 products_catalog.json 的每一筆，找出「品名特徵與分類不符」的品項。

用法：
    python3 scripts/validate_categories.py            # 驗證報告（不改檔）
    python3 scripts/validate_categories.py --fix      # 自動修正高信心項目
    python3 scripts/validate_categories.py --excel 路徑.xlsx   # 輸出重分類草稿 Excel

ERP 匯入新品項後務必跑一次。規則新增直接改下方 RULES。
"""

import json
import re
import sys
import os

CATALOG = os.path.join(os.path.dirname(__file__), '..', 'public', 'products_catalog.json')

# ── 分類規則 ──────────────────────────────────────────────────────
# (品名 regex, 正確分類)。由上而下先中先贏；只在「現分類 != 正確分類」時回報。
RULES = [
    # 假牙 / 塑鋼牙
    (r'^(EFC-[AP]|FX |MILLION|AC 半塑鋼|SIM-P|EFUCERA|Soluute|Crown PX|New Ace|RH 塑鋼|Enigma 塑鋼)', '塑鋼牙'),
    # 氧化鋯塊
    (r'^(Prettau \d|Prettau Zirconia|ZRA[BD]\d|Antomic|Anatomic Coloured|ICE PLUS|ICE ZIRKON)', '氧化鋯塊'),
    # PMMA / 樹脂圓盤
    (r'^(Temp Basic|Temp Premiu|TEMP PREMIUM|MULTISTRATUM|Tecno Med|TECNO MED|ABRO|DENTURE GINGIVA)', 'PMMA 塊'),
    # 金屬圓盤
    (r'^(Sintermetall \d|SINTERNIT|Titan 5|TITANIT|Chrom-cobalt|MEAC\d|MEAL\d)', '金屬材料'),
    # 染液 - 內染（含 SHT Standard / 3D Master 及所有預染液）
    (r'(Color Liquid|Colour Liquid|Color Luquid|Aquarell(?! Set)|Waterbased [A-Z]|Bio-Pigme|Fresco Liquid|內染液)', '染液 - 內染'),
    # 染液 - 外染（烤燒後表面染色：Stain、Artamic）
    (r'(Artamic[\w\s]*Stain|Matchmaker Stain|3D Stain|Initial Spectrum Stain)', '染液 - 外染'),
    # 瓷粉（3D Base 粉末，歸染液色料大分類）
    (r'3D Base [ABCD]|3D Base Glaze', '瓷粉'),
    # 染液試色板
    (r'Colou?r test Plate', '染液 / 色料'),
    # 蠟
    (r'(Curving Wax|Wax White|蠟塊|Wax Disk)', '蠟 / 壓鑄材'),
    # 樹脂材料（義齒床/補修/臨時冠樹脂）
    (r'(Basing Resin|Re-Fine Bright|Ortho Bright|Soft Liner|OSTRON|TEMPSMART|Basis 慢性粉)', '樹脂材料'),
    # CAD/CAM 複合瓷塊
    (r'(CERASMART|CEARSMART)', '玻璃陶瓷'),
    # 器械（American Eagle 刮治器/探針，GC 代理）
    (r'^(AEGA|AEDG|AESM|AEDGM)', '牙科器材'),
    # 植體工具
    (r'^Screwdriver', '植體配件'),
    # 設備零件
    (r'^Spare Part', '設備配件'),
    # 工具
    (r'(瓷粉雕刀|雕刻刀)', '瓷筆 / 刷具'),
    # 染液試色板（用戶指定：Colour test Plate 歸染液/色料）
    (r'Colou?r test Plate', '染液 / 色料'),
    # 比色板
    (r'(Shade Guide|比色板)', '比色板'),
]

# 例外白名單：這些品名雖中規則但分類本來就對（避免誤報）
WHITELIST_CATEGORIES = {
    # pattern 的目標分類本身，以及合理近親
    '塑鋼牙': {'塑鋼牙'},
    '氧化鋯塊': {'氧化鋯塊'},
    'PMMA 塊': {'PMMA 塊'},
    '金屬材料': {'金屬材料'},
    '染液 / 色料': {'染液 / 色料'},
    '染液 - 內染': {'染液 - 內染'},
    '染液 - 外染': {'染液 - 外染'},
    '瓷粉': {'瓷粉'},
    '蠟 / 壓鑄材': {'蠟 / 壓鑄材', '蠟塊'},
    '樹脂材料': {'樹脂材料'},
    '玻璃陶瓷': {'玻璃陶瓷'},
    '牙科器材': {'牙科器材', '工具'},
    '植體配件': {'植體配件'},
    '設備配件': {'設備配件', '3D列印機配件'},
    '瓷筆 / 刷具': {'瓷筆 / 刷具'},
    '比色板': {'比色板'},
}


# ── 品牌前綴規則(2026-07-12 使用者定案:貨號前綴=品牌)─────────────────
# 鍵=貨號開頭的字母段(大寫)。來源:全目錄 6,084 筆實測,每前綴唯一對應一品牌、零衝突。
# 未列入的前綴(VP/ST/C0/BF 等)品牌待使用者確認,不要猜。
BRAND_PREFIXES = {
    'YMH': 'YAMAHACHI', 'ZZ': 'Zirkonzahn', 'BS': '貝施美', 'DSD': 'Davis Schottlander',
    'GC': 'GC / 台灣而至', 'SY': 'Song Young', 'YM': 'YAMAKIN', 'DK': 'DENKEN',
    'SS': 'Sunshine / DR.HOPF', 'CA': 'CAM / 上海穩昊', 'AG': 'ASIGA', 'DUM': 'Dumont',
    'WM': 'WHIP MIX', 'SUN': 'SUN Oberflächentechnik', 'GEN': 'GenCore', 'ADB': 'Prima Dental',
    'KM': 'KO-MAX', 'DT': 'DETAX', 'MT': 'DENTAL ESPAN', 'HD': 'HIGH DENTAL JAPAN',
    'AB': 'Aalba Dent', 'UW': 'URAWA', 'MO': 'MOTYL', 'RD': 'Redon', 'KS': 'KEYSTONE',
    'ME': 'MEDIFIVE', 'MPF': 'MPF', 'CAD': 'CADstar', 'PD': 'PRODENT-HOLLIGER',
    'SA': 'SAEYANG', 'SD': 'Select Dental', 'UG': 'UGin Dental', 'DB': 'DENTBIRD',
    'DKM': 'Dekema', 'PC': 'PACIFIC ABRASIVES', 'PM': 'PROMEDLCA', 'TD': 'Talmax',
    'AF': 'Argofile', 'DOF': 'DOF', 'DP': 'DENKEN', 'OS': 'Olson Saw', 'WF': 'WINFRIED MULLER',
}


def brand_from_code(code: str):
    m = re.match(r'^([A-Za-z]+)', code or '')
    return BRAND_PREFIXES.get(m.group(1).upper()) if m else None


def classify(name: str):
    for pat, cat in RULES:
        if re.search(pat, name, re.I):
            return cat
    return None


def main():
    fix = '--fix' in sys.argv
    excel_path = None
    if '--excel' in sys.argv:
        excel_path = sys.argv[sys.argv.index('--excel') + 1]

    with open(CATALOG, encoding='utf-8') as f:
        catalog = json.load(f)

    mismatches = []
    for p in catalog:
        want = classify(p['name'])
        if want is None:
            continue
        ok_set = WHITELIST_CATEGORIES.get(want, {want})
        if p['category'] in ok_set:
            continue
        mismatches.append((p, want))

    # 品牌檢查:前綴=品牌(空白可 --fix 補;不一致只回報,由人工判斷是 ERP 錯還是新前綴)
    brand_blank, brand_conflict = [], []
    for p in catalog:
        want_b = brand_from_code(p['code'])
        if want_b is None:
            continue
        if not p.get('brand'):
            brand_blank.append((p, want_b))
        elif p['brand'] != want_b:
            brand_conflict.append((p, want_b))

    print(f'總品項: {len(catalog)}，分類疑似錯誤: {len(mismatches)}，'
          f'品牌空白可補: {len(brand_blank)}，品牌與前綴不符: {len(brand_conflict)}')
    for p, want_b in brand_conflict:
        print(f'  品牌不符  {p["code"]}  {p["brand"]!r} → 前綴應為 {want_b!r}')
    from collections import Counter
    stat = Counter(f"{p['category']} → {want}" for p, want in mismatches)
    for k, v in stat.most_common():
        print(f'  {v:>5}  {k}')

    if excel_path:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        wb = Workbook()
        ws = wb.active
        ws.title = '重分類草稿'
        ws.append(['貨號', '品名', '品牌', '原分類', '建議分類', '✏️確認(留空=同意/填別的分類)'])
        hdr = PatternFill('solid', start_color='1F4E79')
        for c in ws[1]:
            c.fill = hdr
            c.font = Font(bold=True, color='FFFFFF', name='Arial', size=10)
            c.alignment = Alignment(horizontal='center')
        edit = PatternFill('solid', start_color='FFF2CC')
        for p, want in mismatches:
            ws.append([p['code'], p['name'], p['brand'], p['category'], want, ''])
            ws.cell(row=ws.max_row, column=6).fill = edit
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = Font(name='Arial', size=10)
        for i, w in enumerate([22, 48, 16, 14, 14, 24], 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = 'A2'
        ws.auto_filter.ref = f'A1:F{ws.max_row}'
        wb.save(excel_path)
        print(f'草稿已輸出: {excel_path}')

    if fix:
        for p, want in mismatches:
            p['category'] = want
        for p, want_b in brand_blank:  # 只補空白,不覆蓋既有品牌
            p['brand'] = want_b
        with open(CATALOG, 'w', encoding='utf-8') as f:
            json.dump(catalog, f, ensure_ascii=False, indent=1)
        print(f'已修正 {len(mismatches)} 筆分類、補 {len(brand_blank)} 筆空白品牌')


if __name__ == '__main__':
    main()
