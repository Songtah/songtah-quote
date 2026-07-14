#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
新分類遷移:松達產品分類總表(2026-07-14)→ products_catalog.json + data/product-taxonomy.json

- SKU 對 SKU 改寫 mainCategory(11 主分類)/category(62 功能分類)/productType(8 商品型態),
  新增 mainCategoryId/categoryId(英文 slug)、seriesName(總表系列)、needsReview(待覆核旗標)。
- 同時從總表字典頁產出 data/product-taxonomy.json(id/名稱/父子關係),供 validator 與 UI 共用。
- 用法:python3 scripts/migrate-taxonomy.py <總表.xlsx>            → dry-run 差異報告
        python3 scripts/migrate-taxonomy.py <總表.xlsx> --execute  → 真寫(git 版控可回溯)
"""
import json, sys, os
from collections import Counter
from openpyxl import load_workbook

ROOT = os.path.join(os.path.dirname(__file__), '..')
CATALOG = os.path.join(ROOT, 'public', 'products_catalog.json')
TAXONOMY_OUT = os.path.join(ROOT, 'data', 'product-taxonomy.json')

if len(sys.argv) < 2:
    print('用法: python3 scripts/migrate-taxonomy.py <總表.xlsx> [--execute]'); sys.exit(1)
XLSX = sys.argv[1]
EXECUTE = '--execute' in sys.argv

wb = load_workbook(XLSX, data_only=True)

# ── 讀字典頁 → taxonomy.json ────────────────────────────────────────
ws = wb['分類字典']
mains, funcs, forms = [], [], []
seen_m, seen_f = set(), set()
forms_started = forms_done = False
for row in ws.iter_rows(min_row=6, values_only=True):
    if row[0] and row[1] and row[0] not in seen_m:
        mains.append({'id': str(row[0]).strip(), 'name': str(row[1]).strip()}); seen_m.add(row[0])
    if row[3] and row[4] and row[3] not in seen_f:
        funcs.append({'id': str(row[3]).strip(), 'name': str(row[4]).strip(),
                      'mainId': str(row[5] or '').strip(), 'mainName': str(row[6] or '').strip()})
        seen_f.add(row[3])
    # 商品型態:只取第一段連續區塊(字典頁下方併了 3D 列印補充表,遇空列即停避免混入)
    if not forms_done:
        if row[8] and row[9]:
            forms.append({'id': str(row[8]).strip(), 'name': str(row[9]).strip()}); forms_started = True
        elif forms_started:
            forms_done = True

# ── 讀商品分類頁 → SKU 映射 ────────────────────────────────────────
ws = wb['商品分類']
hdr = [c.value for c in ws[1]]
idx = {h: i for i, h in enumerate(hdr)}
newmap = {}
for row in ws.iter_rows(min_row=2, values_only=True):
    if not row[0]: continue
    sku = str(row[0]).strip()
    newmap[sku] = {
        'mainCategoryId': str(row[idx['主分類 ID']] or '').strip(),
        'mainCategory':   str(row[idx['主分類']] or '').strip(),
        'categoryId':     str(row[idx['功能分類 ID']] or '').strip(),
        'category':       str(row[idx['功能分類']] or '').strip(),
        'productType':    str(row[idx['商品型態']] or '').strip(),
        'seriesName':     str(row[idx['系列名稱']] or '').strip(),
        'needsReview':    str(row[idx['需覆核']] or '').strip() == '是' or str(row[idx['分類狀態']] or '').strip() in ('待覆核', '未判定'),
    }

# ── 套到 catalog ───────────────────────────────────────────────────
with open(CATALOG, encoding='utf-8') as f:
    catalog = json.load(f)

chg = Counter(); miss = []
for p in catalog:
    m = newmap.get(p['code'])
    if not m:
        miss.append(p['code']); continue
    for k in ('category', 'mainCategory', 'productType'):
        if (p.get(k) or '') != m[k]: chg[k] += 1
    p['category'] = m['category']
    p['mainCategory'] = m['mainCategory']
    p['productType'] = m['productType']
    p['mainCategoryId'] = m['mainCategoryId']
    p['categoryId'] = m['categoryId']
    if m['seriesName']: p['seriesName'] = m['seriesName']
    else: p.pop('seriesName', None)
    if m['needsReview']: p['needsReview'] = True
    else: p.pop('needsReview', None)

print(f'catalog {len(catalog)} 筆;總表映射 {len(newmap)} 筆;缺漏 {len(miss)}{miss[:5] if miss else ""}')
print(f'變動:category {chg["category"]} / mainCategory {chg["mainCategory"]} / productType {chg["productType"]}')
print(f'系列 {sum(1 for p in catalog if p.get("seriesName"))} 筆;需覆核 {sum(1 for p in catalog if p.get("needsReview"))} 筆')
print(f'字典:主分類 {len(mains)}、功能分類 {len(funcs)}、商品型態 {len(forms)}({", ".join(f["name"] for f in forms)})')

# 一致性驗證:SKU 頁實際值必須都在字典裡(防字典讀取錯位)
dict_forms = {f['name'] for f in forms}
dict_funcs = {f['name'] for f in funcs}
dict_mains = {m['name'] for m in mains}
bad_form = {m['productType'] for m in newmap.values() if m['productType'] and m['productType'] not in dict_forms}
bad_func = {m['category'] for m in newmap.values() if m['category'] and m['category'] not in dict_funcs}
bad_main = {m['mainCategory'] for m in newmap.values() if m['mainCategory'] and m['mainCategory'] not in dict_mains}
if bad_form or bad_func or bad_main:
    print(f'⚠ SKU 頁有字典外的值 → 型態:{bad_form or "-"} 功能:{bad_func or "-"} 主:{bad_main or "-"}')
    sys.exit(1)
print('✓ SKU 頁所有值皆在字典內')

if not EXECUTE:
    print('\n(DRY-RUN,未寫入。加 --execute 真寫)'); sys.exit(0)

taxonomy = {
    '_說明': '產品分類字典(源自 松達產品分類總表 2026-07-14)。三層:主分類→功能分類(含所屬主分類)+商品型態。validator 與 UI 共用;改分類先改總表再跑 scripts/migrate-taxonomy.py。',
    'version': '2026-07-14',
    'mainCategories': mains,
    'funcCategories': funcs,
    'productForms': forms,
}
os.makedirs(os.path.dirname(TAXONOMY_OUT), exist_ok=True)
with open(TAXONOMY_OUT, 'w', encoding='utf-8') as f:
    json.dump(taxonomy, f, ensure_ascii=False, indent=1)
with open(CATALOG, 'w', encoding='utf-8') as f:
    json.dump(catalog, f, ensure_ascii=False, indent=1)
print(f'\n✅ 已寫入 {CATALOG} 與 {TAXONOMY_OUT}')
