#!/usr/bin/env python3
"""
將「產品價格主檔.xlsx」的售價/優惠價/技術規格合併回 public/products_catalog.json。

用法：
    python3 scripts/merge_catalog_prices.py [Excel路徑]

預設讀取 ~/Desktop/Songtah/產品價格主檔.xlsx。
以貨號為對照鍵；Excel 中留空的欄位不會清除 JSON 既有值。
執行後輸出變動統計，git diff 可檢視所有變更。
"""

import json
import sys
import os

from openpyxl import load_workbook

EXCEL_PATH = sys.argv[1] if len(sys.argv) > 1 else os.path.expanduser(
    '~/Desktop/Songtah/產品價格主檔.xlsx'
)
JSON_PATH = os.path.join(os.path.dirname(__file__), '..', 'public', 'products_catalog.json')


def main():
    wb = load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    ws = wb.active

    # 讀 Excel：貨號 → {price, salePrice, spec}
    updates = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        code = str(row[0]).strip() if row[0] else ''
        if not code:
            continue
        price, sale, spec = row[5], row[6], row[7]
        entry = {}
        if price is not None and str(price).strip() != '':
            entry['price'] = float(price)
        if sale is not None and str(sale).strip() != '':
            entry['salePrice'] = float(sale)
        if spec is not None and str(spec).strip() != '':
            entry['spec'] = str(spec).strip()
        if entry:
            updates[code] = entry

    with open(JSON_PATH, encoding='utf-8') as f:
        catalog = json.load(f)

    changed = 0
    not_found = set(updates.keys())
    for p in catalog:
        u = updates.get(p['code'])
        if not u:
            continue
        not_found.discard(p['code'])
        before = (p.get('price'), p.get('salePrice'), p.get('spec'))
        p.update(u)
        after = (p.get('price'), p.get('salePrice'), p.get('spec'))
        if before != after:
            changed += 1

    with open(JSON_PATH, 'w', encoding='utf-8') as f:
        json.dump(catalog, f, ensure_ascii=False, indent=1)

    print(f'Excel 有資料的品項：{len(updates)}')
    print(f'實際變動：{changed} 筆')
    if not_found:
        print(f'⚠️ Excel 有但 JSON 找不到的貨號（{len(not_found)} 個）：')
        for c in sorted(not_found)[:20]:
            print(f'   {c}')


if __name__ == '__main__':
    main()
