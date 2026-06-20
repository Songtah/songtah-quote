#!/usr/bin/env python3
"""
建立學校靜態參照 data/schools.json（教育部各級學校名錄）。

學校極少變動，故當「靜態參照」用：一學年更新一次，手動跑此腳本即可。
來源：教育部統計處 各級學校名錄 https://stats.moe.gov.tw/files/school/<學年>/...

對崧達（牙材）相關的學校層級：大專校院（牙體技術系）、高級中等學校、大專附設高職部。
產出格式：{ "<學校代碼>": { "name", "city", "address", "kind" } }

用法：python3 scripts/build_schools.py [學年，預設 114]
"""
import sys, json, io, ssl, urllib.request, re
from openpyxl import load_workbook

# MOE 伺服器憑證鏈為自簽，手動建檔腳本抓公開政府資料 → 略過驗證
SSL_CTX = ssl.create_default_context()
SSL_CTX.check_hostname = False
SSL_CTX.verify_mode = ssl.CERT_NONE

YEAR = sys.argv[1] if len(sys.argv) > 1 else '114'
BASE = f'https://stats.moe.gov.tw/files/school/{YEAR}'

# 檔名 → 層級標籤（崧達牙材相關層級）
FILES = {
    'u1_new.xlsx': '大專校院',
    'high.xlsx':   '高級中等學校',
    'highA.xlsx':  '大專附設高職部',
    'highT.xlsx':  '大專附設進修學校',
}

def fetch_xlsx(url):
    req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
    with urllib.request.urlopen(req, timeout=60, context=SSL_CTX) as r:
        return io.BytesIO(r.read())

def find_header(rows):
    """找含『代碼』『學校名稱』『地址』的表頭列，回 (idx, colmap)。"""
    for i, row in enumerate(rows[:8]):
        vals = [str(c).strip() if c is not None else '' for c in row]
        if any(v == '代碼' or v.endswith('代碼') for v in vals) and any('學校名稱' in v for v in vals):
            colmap = {}
            for j, v in enumerate(vals):
                if v.endswith('代碼') or v == '代碼': colmap['code'] = j
                elif '學校名稱' in v:                  colmap['name'] = j
                elif v == '縣市名稱' or v == '縣市':    colmap['city'] = j
                elif v == '地址':                       colmap['address'] = j
            if 'code' in colmap and 'name' in colmap and 'address' in colmap:
                return i, colmap
    return None, None

def clean_addr(a):
    # 去除地址開頭的 [郵遞區號]
    return re.sub(r'^\[\d+\]', '', (a or '').strip()).strip()

schools = {}
for fname, kind in FILES.items():
    url = f'{BASE}/{fname}'
    try:
        wb = load_workbook(fetch_xlsx(url), read_only=True, data_only=True)
    except Exception as e:
        print(f'⚠ 略過 {fname}: {e}')
        continue
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    hi, cm = find_header(rows)
    if hi is None:
        print(f'⚠ {fname} 找不到表頭，略過')
        continue
    n = 0
    for row in rows[hi + 1:]:
        code = row[cm['code']]
        name = row[cm['name']]
        if not code or not name: continue
        code = str(code).strip()
        if not re.match(r'^[0-9A-Za-z]{3,}$', code): continue
        addr = clean_addr(str(row[cm['address']]) if row[cm['address']] else '')
        # 縣市欄常是代碼（如 [38]），改由地址開頭解析
        cm_city = re.match(r'^(.*?[市縣])', addr)
        city = cm_city.group(1) if cm_city else ''
        schools[code] = { 'name': str(name).strip(), 'city': city, 'address': addr, 'kind': kind }
        n += 1
    print(f'  {kind}（{fname}）：{n} 校')

with open('data/schools.json', 'w', encoding='utf-8') as f:
    json.dump({ 'year': YEAR, 'schools': schools }, f, ensure_ascii=False, indent=1)
print(f'\n✅ 共 {len(schools)} 校 → data/schools.json')
